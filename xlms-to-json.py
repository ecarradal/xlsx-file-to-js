import openpyxl
from pathlib import Path

xlsx_file = Path('filename.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

# privileges obj
objPrivileges = '''
    email: "{}",
    actives: {},
    process: {},
    providers: {},
    vacancies: {},
    MEX: {},
    COL: {},
    CHL: {},
    PER: {},
'''

def check_value(value):
  if(value):
    return "true"
  else:
    return "false"

col_names = []
with open("privilegesFile.js", "r+") as f:
    old = f.read()
    f.write("[\n")

    for row in sheet.iter_rows(2, sheet.max_row):
        if(row[0].value != None):
            col_names.append(row[0].value)
            f.write("\t{")
            f.write(objPrivileges.format(
              row[0].value,
              check_value(row[1].value),
              check_value(row[2].value),
              check_value(row[3].value),
              check_value(row[4].value),
              check_value(row[5].value),
              check_value(row[6].value),
              check_value(row[7].value),
              check_value(row[8].value)
            ))
            f.write("\t},\n")

    f.write("]\n")
    print(col_names)
